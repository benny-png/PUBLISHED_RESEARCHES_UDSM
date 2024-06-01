import generated_resource
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Create a new workbook
wb = Workbook()

# Define the sheet names and headers for each quarter
quarters = {
    "Q1": "1 January - 31 March 2024",
    "Q2": "1 April - 30 June 2024",
    "Q3": "1 July - 30 September 2024",
    "Q4": "1 October - 31 December 2024"
}

# Define the column headers
columns = [
    "Department", 
    " SN ",  # Add one space before and after "SN"
    "Journal articles", 
    "",  # Placeholder for the jump after Journal articles
    "",  # Placeholder for the jump after Journal articles
    "Books", 
    "",  # Placeholder for the jump after Books
    "",  # Placeholder for the jump after Books
    "Conference",
    "",  # Placeholder for the jump after Conference
    "",  # Placeholder for the jump after Conference
    "Intellectual Property Rights (IPR)"
]

# Define the sub-column headers for "Journal articles"
journal_sub_columns = [
    "Title of article",
    "Authors",
    "Journal",
    "Publisher",
    "Publication Date"
]

# Define the sub-column headers for "Books"
books_sub_columns = [
    "Title of book",
    "Authors",
    "Publisher",
    "Date of publication",
    "Funder"
]

# Define the sub-column headers for "Conference"
conference_sub_columns = [
    "Title of paper",
    "Authors",
    "Name of conference",
    "Organizer",
    "Place of conference",
    "Date of conference",
    "Funder"
]

# Define the sub-column headers for "Intellectual Property Rights (IPR)"
ipr_sub_columns = [
    "Type of IPR (Patent, Copyright, Trademark)",
    "Title of IPR",
    "Individuals involved",
    "IPR Organization",
    "Status (Granted/Filed)",
    "Date of Award/Submission",
    "Scope"
]

# Define the departments
departments = ["ETE", "CSE"]

# Add sheets for each quarter, set the headers, and add the header row
for quarter, header in quarters.items():
    ws = wb.create_sheet(title=quarter)
    
    # Column number tracker
    col_num = 1
    
    # Add the main column headers
    for column_title in columns:
        if column_title == "Journal articles":
            # Merge cells for the "Journal articles" header
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(journal_sub_columns) - 1)
            ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
            # Add sub-column headers for "Journal articles"
            for sub_col_num, sub_col_title in enumerate(journal_sub_columns, start=col_num):
                ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
            col_num += len(journal_sub_columns)
        elif column_title == "Books":
            # Merge cells for the "Books" header
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(books_sub_columns) - 1)
            ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
            # Add sub-column headers for "Books"
            for sub_col_num, sub_col_title in enumerate(books_sub_columns, start=col_num):
                ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
            col_num += len(books_sub_columns)
        elif column_title == "Conference":
            # Merge cells for the "Conference" header
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(conference_sub_columns) - 1)
            ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
            # Add sub-column headers for "Conference"
            for sub_col_num, sub_col_title in enumerate(conference_sub_columns, start=col_num):
                ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
            col_num += len(conference_sub_columns)
        elif column_title == "Intellectual Property Rights (IPR)":
            # Merge cells for the "Intellectual Property Rights (IPR)" header
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(ipr_sub_columns) - 1)
            ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
            # Add sub-column headers for "Intellectual Property Rights (IPR)"
            for sub_col_num, sub_col_title in enumerate(ipr_sub_columns, start=col_num):
                ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
            col_num += len(ipr_sub_columns)
        elif column_title in ["", "Intellectual Property Rights (IPR)"]:
            # Skip columns for the jumps and add headers for the remaining columns
            if column_title != "":
                ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title.strip()), 12)  # Minimum width is 12
            col_num += 1
        else:
            ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
            # Set column width based on the length of the column name
            if column_title.strip() == "SN":  # Strip to remove extra spaces
                ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title), 3)  # Adjust to fit "SN"
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title.strip()), 12)  # Minimum width is 12
            col_num += 1

    # Add rows for ETE and CSE departments
    current_row = 3  # Start from the third row since first two are headers
    for department in departments:
        ws.cell(row=current_row, column=1, value=department)
        current_row += 1

# Remove the default sheet created with the workbook
if 'Sheet' in wb.sheetnames:
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

# Save the workbook to a file named 'COICT.xlsx'
file_name = "COICT.xlsx"
wb.save(file_name)

print(f"Workbook saved as '{file_name}'")

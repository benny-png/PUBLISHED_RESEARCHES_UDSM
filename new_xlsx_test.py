import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd

def export_to_excel(csv_file_path, output_file_name="COICT.xlsx"):
    # Read the CSV file
    df = pd.read_csv(csv_file_path)

    # Create a new workbook
    wb = Workbook()

    # Define the sheet names and headers for each quarter
    quarters = {
        "Q1": "1 January - 31 March 2024",
        "Q2": "1 April - 30 June 2024",
        "Q3": "1 July - 30 September 2024",
        "Q4": "1 October - 31 December 2024"
    }

    # Define the column headers
    columns = [
        "DEPARTMENT", 
        " SN ",  # Add one space before and after "SN"
        "JOURNAL ARTICLES", 
        "",  # Placeholder for the jump after Journal articles
        "",  # Placeholder for the jump after Journal articles
        "BOOKS", 
        "",  # Placeholder for the jump after Books
        "",  # Placeholder for the jump after Books
        "CONFERENCE", 
        "INTELLECTUAL PROPERTY RIGHTS (IPR)"
    ]

    # Define the sub-column headers for "Journal articles"
    journal_sub_columns = [
        "TITLE OF ARTICLE",
        "AUTHORS",
        "JOURNAL",
        "PUBLISHER",
        "PUBLICATION DATE"
    ]

    # Define the sub-column headers for "Books"
    books_sub_columns = [
        "TITLE OF BOOK",
        "AUTHORS",
        "PUBLISHER",
        "DATE OF PUBLICATION",
        "FUNDER"
    ]

    # Define the sub-column headers for "Conference"
    conference_sub_columns = [
        "TITLE OF PAPER",
        "AUTHORS",
        "NAME OF CONFERENCE",
        "ORGANIZER",
        "PLACE OF CONFERENCE",
        "DATE OF CONFERENCE",
        "FUNDER"
    ]

    # Define the departments
    departments = ["ETE", "CSE"]

    for quarter, header in quarters.items():
        ws = wb.create_sheet(title=quarter)
        
        # Column number tracker
        col_num = 1
        
        # Add the main column headers
        for column_title in columns:
            if column_title == "JOURNAL ARTICLES":
                # Merge cells for the "Journal articles" header
                ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(journal_sub_columns) - 1)
                ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
                # Add sub-column headers for "Journal articles"
                for sub_col_num, sub_col_title in enumerate(journal_sub_columns, start=col_num):
                    ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                    ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
                col_num += len(journal_sub_columns)
            elif column_title == "BOOKS":
                # Merge cells for the "Books" header
                ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(books_sub_columns) - 1)
                ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
                # Add sub-column headers for "Books"
                for sub_col_num, sub_col_title in enumerate(books_sub_columns, start=col_num):
                    ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                    ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
                col_num += len(books_sub_columns)
            elif column_title == "CONFERENCE":
                # Merge cells for the "Conference" header
                ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(conference_sub_columns) - 1)
                ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
                # Add sub-column headers for "Conference"
                for sub_col_num, sub_col_title in enumerate(conference_sub_columns, start=col_num):
                    ws.cell(row=2, column=sub_col_num, value=sub_col_title).alignment = Alignment(horizontal='center')
                    ws.column_dimensions[get_column_letter(sub_col_num)].width = max(len(sub_col_title.strip()), 12)  # Minimum width is 12
                col_num += len(conference_sub_columns)
            elif column_title in ["", "INTELLECTUAL PROPERTY RIGHTS (IPR)"]:
                # Skip columns for the jumps and add headers for the remaining columns
                if column_title != "":
                    ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
                    ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title.strip()), 12)  # Minimum width is 12
                col_num += 1
            else:
                ws.cell(row=1, column=col_num, value=column_title).alignment = Alignment(horizontal='center')
                # Set column width based on the length of the column name
                if column_title.strip() == "SN":  # Strip to remove extra spaces
                    ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title), 4)  # Adjust to fit "SN"
                else:
                    ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title.strip()), 12)  # Minimum width is 12
                col_num += 1


     
        # Add rows for ETE and CSE departments
        current_row = 3  # Start from the third row since first two are headers
        for department in departments:
            ws.cell(row=current_row, column=1, value=department)
            current_row += 1

    # Remove the default sheet created with the workbook
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

    # Map the data from the CSV file to the Excel workbook
    max_widths = {get_column_letter(i): 0 for i in range(1, len(columns) + 1)}

    for index, row in df.iterrows():
        # Get the details from the CSV
        authors = row['AUTHORS']
        year = row['YEAR']
        title = row['TITLE']
        journal = row['JOURNAL']

        # Find the appropriate sheet and row for the current data
        # Assuming all data belongs to Q1 for this example
        ws = wb['Q1']
        
        # Find the next available row for the journal articles section
        for row_num in range(3, ws.max_row + 2):
            if ws.cell(row=row_num, column=3).value is None:
                ws.cell(row=row_num, column=3, value=title)
                ws.cell(row=row_num, column=4, value=authors)
                ws.cell(row=row_num, column=5, value=journal)
                ws.cell(row=row_num, column=6, value="-")  # Publisher not provided in CSV
                ws.cell(row=row_num, column=7, value=year)

                # Update the maximum widths for columns
                max_widths[get_column_letter(3)] = max(max_widths[get_column_letter(3)], len(str(title)))
                max_widths[get_column_letter(4)] = max(max_widths[get_column_letter(4)], len(str(authors)))
                max_widths[get_column_letter(5)] = max(max_widths[get_column_letter(5)], len(str(journal)))
                max_widths[get_column_letter(6)] = max(max_widths[get_column_letter(6)], len("-"))
                max_widths[get_column_letter(7)] = max(max_widths[get_column_letter(7)], len(str(year)))
                break


# Adjust column widths based on the maximum length of data or header
    for col_letter, max_width in max_widths.items():
        column_index = openpyxl.utils.column_index_from_string(col_letter) - 1
        header_length = len(columns[column_index].strip()) if columns[column_index].strip() != "" else 20
        ws.column_dimensions[col_letter].width = max(max_width, header_length)

   # Save the workbook to a file
    wb.save(output_file_name)

    print(f"Workbook saved as '{output_file_name}'")

# Example usage:
# export_to_excel("Research_paper_details.csv", "COICT.xlsx")
